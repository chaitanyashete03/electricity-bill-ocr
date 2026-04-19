import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import hashlib
from datetime import datetime
import os

class ExcelEngine:
    def __init__(self):
        pass

    def generate_dynamic_excel(self, consumer_data: dict, output_path: str):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "E-Bill Analysis"
        
        # 1. Styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="4F81BD")
        bold_font = Font(bold=True)
        center_align = Alignment(horizontal="center", vertical="center")
        
        # Column widths
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        
        # 2. Consumer Metadata (Rows 3-7)
        ws["B3"] = "Consumer Name:"
        ws["B4"] = "Consumer No:"
        ws["B5"] = "Fixed Charges (₹):"
        ws["B6"] = "Sanctioned Load:"
        ws["B7"] = "Solar Panel Input (W):"
        
        for r in range(3, 8):
            ws[f"B{r}"].font = bold_font
            
        ws["D3"] = consumer_data.get("consumer_name", {}).get("value", "")
        ws["D4"] = consumer_data.get("consumer_number", {}).get("value", "")
        
        try:
            fc = float(consumer_data.get("fixed_charges", {}).get("value", 0))
        except:
            fc = 0.0
        ws["D5"] = fc
        
        ws["D6"] = consumer_data.get("sanctioned_load", {}).get("value", "")
        ws["C7"] = 600 # Default user input for solar panel wattage
        ws["C7"].font = Font(bold=True, color="0070C0")
        
        # 3. Main Data Headers (Row 8)
        headers = [("B8", "Sr.No"), ("C8", "Month"), ("D8", "Units"), ("E8", "Bill Amount"), ("F8", "Unit Cost")]
        for cell_ref, text in headers:
            ws[cell_ref] = text
            ws[cell_ref].font = header_font
            ws[cell_ref].fill = header_fill
            ws[cell_ref].alignment = center_align
            
        # 4. Insert dynamic extracted row at Row 9
        ws["B9"] = 1
        ws["C9"] = consumer_data.get("bill_month", {}).get("value", "")
        try:
            units = float(consumer_data.get("units", {}).get("value", 0))
        except:
            units = 0.0
            
        try:
            bill_amount = float(consumer_data.get("bill_amount", {}).get("value", 0))
        except:
            bill_amount = 0.0
            
        ws["D9"] = units
        ws["E9"] = bill_amount
        ws["F9"] = "=(E9-D$5)/D9"
        
        # Currency formatting
        ws["E9"].number_format = '₹#,##0.00'
        ws["F9"].number_format = '₹#,##0.00'
        
        # 5. Summary Formulas (Rows 22-28)
        ws["B22"] = "Average Units:"
        ws["C22"] = "=AVERAGE(D9:D21)"
        
        ws["B23"] = "Required kW:"
        ws["C23"] = "=(C22*12*1.1)/1400"
        
        ws["B24"] = "Solar Panels Needed:"
        ws["C24"] = "=(C23/$C$7)*1000"
        
        ws["B25"] = "Solar Capacity (kW):"
        ws["C25"] = "=ROUND(C24,0)*$C$7/1000"
        
        for r in range(22, 26):
            ws[f"B{r}"].font = bold_font
            
        # Save fresh workbook
        wb.save(output_path)
        return True
